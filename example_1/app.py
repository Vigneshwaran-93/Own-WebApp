from flask import Flask, render_template, request, redirect, url_for
import openpyxl

app = Flask(__name__)

# Create or load an existing Excel workbook to store contact data
def load_workbook():
    try:
        wb = openpyxl.load_workbook('contact_data.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Name', 'Email', 'Message'])  # Add headers to the first row
        wb.save('contact_data.xlsx')
    return wb

@app.route('/')
def index():
    return render_template('contact.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']
    message = request.form['message']
    
    # Store the contact data in the Excel file
    wb = load_workbook()
    sheet = wb.active
    sheet.append([name, email, message])
    wb.save('contact_data.xlsx')
    
    return 'Data stored successfully! Thank you for contacting us.'

if __name__ == '__main__':
    app.run(debug=True)
